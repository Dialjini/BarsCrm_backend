# -*- coding: utf-8 -*-

import openpyxl

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
            'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
            'W', 'X', 'Y', 'Z']

def getLetter(num):
    letter = ''
    while int(num) > 0:
        letter += alphabet[(int(num) % 26 - 1)]
        num /= 26

    return letter[::-1].replace('AZ', 'Z')

def firstScenario(data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'].value = 'Товар'
    sheet['B1'].value = 'Клиент'
    sheet['C1'].value = 'Объем, кг.'
    sheet['D1'].value = 'Цена, руб.'
    sheet['E1'].value = 'Доставка'
    sheet['F1'].value = 'Привет'
    sheet['G1'].value = 'Себестоимость'
    sheet['H1'].value = 'Закупочная цена'
    sheet['I1'].value = 'Заработали'
    sheet['J1'].value = 'Прибыль'
    column_counter = 1
    for i in data:
        column_counter += 1
        for j in range(10):
            sheet[alphabet[j] + str(column_counter)].value = list(i.values())[j]

    wb.save('app/upload/last_stat.xlsx')

    return '/upload'

def secondScenario(data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'].value = 'Товар'
    counter = 1
    column_counter = 1
    sum_count = []
    for i in data[0]['list']:
        counter += 1
        sheet[getLetter(counter) + '1'].value = i['name']
        sum_count.append(0)
    counter = 1
    for i in data:
        column_counter += 1
        sheet[getLetter(counter) + str(column_counter)].value = i['name']
        for item in i['list']:
            counter += 1
            sheet[getLetter(counter) + str(column_counter)].value = item['volume']
            sum_count[counter - 2] += int(item['volume'])
        counter = 1
    for i in sum_count:
        counter += 1
        sheet[getLetter(counter) + str(column_counter + 1)].value = str(i)
    wb.save('app/upload/last_stat.xlsx')

    return '/upload'

def thirdScenario(data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'].value = 'Клиент'
    sheet['B1'].value = 'Товары'
    sheet['C1'].value = 'Вес'
    sheet['E1'].value = 'Сумма, руб.'

    column_counter = 1
    for i in data:
        column_counter += 1
        for j in range(4):
            sheet[alphabet[j] + str(column_counter)].value = list(i.values())[j]

    wb.save('app/upload/last_stat.xlsx')
    return '/upload'

def fourthScenario(data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'].value = 'Клиент'
    sheet['B1'].value = 'Объем'
    sheet['C1'].value = 'Сколько'
    sheet['D1'].value = 'Сумма, руб.'
    sheet['E1'].value = 'Итого'

    column_counter = 1
    for i in data:
        column_counter += 1
        for j in range(5):
            sheet[alphabet[j] + str(column_counter)].value = list(i.values())[j]

    wb.save('app/upload/last_stat.xlsx')
    return '/upload'

def fifthScenario(data):
    print(data)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'].value = 'Менеджер'
    counter = 1
    column_counter = 1
    sum_count = []
    for i in data[0]['list']:
        counter += 1
        if i['name'] == 'total':
            i['name'] = 'Итого'
        sheet[getLetter(counter) + '1'].value = i['name']
        sum_count.append(0)
    counter = 1
    for i in data:
        column_counter += 1
        sheet[getLetter(counter) + str(column_counter)].value = i['manager']
        for item in i['list']:
            counter += 1
            print('------------------------test_5-th_scenario-----------------------')
            print(item)
            sheet[getLetter(counter) + str(column_counter)].value = item['volume']
            sum_count[counter - 2] += int(item['volume'])
        counter = 1
    for i in sum_count:
        counter += 1
        sheet[getLetter(counter) + str(column_counter + 1)].value = str(i)

    wb.save('app/upload/last_stat.xlsx')
    return '/upload'

def sixthScenario(data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'].value = 'Менеджер'
    sheet['B1'].value = 'Наименование'
    sheet['C1'].value = 'Дата последнего комментария'

    column_counter = 1
    for i in data:
        column_counter += 1
        for j in range(3):
            sheet[alphabet[j] + str(column_counter)].value = list(i.values())[j]

    wb.save('app/upload/last_stat.xlsx')
    return '/upload'

def seventhScenario(data):
    print(data)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'].value = 'Менеджер'
    sheet['B1'].value = 'Дата'
    sheet['C1'].value = 'Наименование'
    sheet['D1'].value = 'Товар'
    sheet['E1'].value = 'Объем'
    sheet['F1'].value = 'Бонус'

    man_prev = data[0]['manager']
    column_counter = 2
    bonus = ''
    item = ''
    client = ''
    for i in data:
        if i['manager'] == man_prev:
            bonus = i['amount_bonus']
            item = i['items_mound']
            client = i['clients']
        else:
            man_prev = i['manager']
            sheet[alphabet[6] + str(column_counter - 1)].value = 'Кол-во новых клиентов: ' + str(client)
            sheet[alphabet[7] + str(column_counter - 1)].value = 'Кол-во товара категории "Насыпь": ' + str(item)
            sheet[alphabet[8] + str(column_counter - 1)].value = 'Итого: ' + str(bonus)
        for j in range(6):
            sheet[alphabet[j] + str(column_counter)].value = list(i.values())[j]
        column_counter += 1
    sheet[alphabet[6] + str(column_counter - 1)].value = 'Новых клиентов: ' + str(client)
    sheet[alphabet[7] + str(column_counter - 1)].value = 'Насыпь: ' + str(item)
    sheet[alphabet[8] + str(column_counter - 1)].value = 'Бонус за месяц: ' + str(bonus)


    wb.save('app/upload/last_stat.xlsx')
    return '/upload'


def createExel(id, data):
    if int(id) == 0:
        return firstScenario(data)
    if int(id) == 1:
        return secondScenario(data)
    if int(id) == 2:
        return thirdScenario(data)
    if int(id) == 3:
        return fourthScenario(data)
    if int(id) == 4:
        return fifthScenario(data)
    if int(id) == 5:
        return sixthScenario(data)
    if int(id) == 6:
        return seventhScenario(data)

    return 'ok'

