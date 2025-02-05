from app import models, db
import json
import openpyxl


# -------------------------------------------------------Roles----------------------------------------------------------
def add_roles():
    role_list = ['Собственник', 'Директор', 'Генеральный директор', 'Заместитель директора',
                 'Председатель', 'Главный бухгалтер','Бухгалтер', 'Снабжение', 'Зоотехник', 'Агроном', 'Секретарь',
                 'Логист', 'Зав. гаражом', 'Водитель']
    for i in range(0, 14):
        role = models.Role()
        role.Name = role_list[i]
        role.Priority = int(i)
        db.session.add(role)
    db.session.commit()
# -------------------------------------------------------User-----------------------------------------------------------
def add_user(login, password, email, role, name, avatar='default'):
    table = models.User()
    table.login = login
    table.password = password
    table.email = email
    table.role = role
    table.avatar = avatar
    table.name = name

    db.session.add(table)
    db.session.commit()
# -------------------------------------------------------Client---------------------------------------------------------

def get_clients_contacts(sheet, parent_table, col, client_id):
    i = col + 1
    print(4)
    while(sheet['A' + str(i)].value == None):
        table = models.Contacts()
        table.Position = sheet['B' + str(i)].value
        table.Name = sheet['C' + str(i)].value
        table.Number = sheet['D' + str(i)].value
        table.Email = sheet['E' + str(i)].value
        table.Comment = sheet['F' + str(i)].value
        table.Department = sheet['G' + str(i)].value
        table.Group = sheet['H' + str(i)].value
        table.Manager = sheet['I' + str(i)].value
        table.Date = sheet['J' + str(i)].value
        table.Birthday = sheet['K' + str(i)].value
        table.Owner = parent_table
        table.Client_id = client_id
        i = i + 1
        parent_table.Contacts.append(table)
        if (i - 1) == sheet.max_row:
            return parent_table.Contacts
    return parent_table.Contacts

def get_clients_notes(sheet, parent_table, col, client_id):
    print(3)
    i = col + 1
    while(sheet['A' + str(i)].value == None):
        table = models.Notes()
        if sheet['B' + str(i)].value == '+':
            table.Done = True
        else:
            table.Done = False

        table.Type = sheet['C' + str(i)].value
        table.Date = sheet['D' + str(i)].value
        table.Note = sheet['E' + str(i)].value
        table.Manager = sheet['F' + str(i)].value
        table.File_Path = sheet['H' + str(i)].value
        table.Author = parent_table
        table.Client_id = client_id
        i = i + 1
        parent_table.Notes.append(table)

        if (i - 1) == sheet.max_row:
            return {'notes': parent_table.Notes, 'contacts': parent_table.Contacts}

    if sheet['A' + str(i)].value == 'Контактные лица':
        parent_table.Contacts = get_clients_contacts(sheet=sheet, parent_table=parent_table, col=i, client_id=client_id)

    return {'notes': parent_table.Notes, 'contacts': parent_table.Contacts}

def get_client_from_xlsx(sheet, parent_table, col=1):
    client_id = 1
    print(2)
    for i in range(col, sheet.max_row):
        if(sheet['A'+str(i)].value == 'Компания'):
            table = parent_table()
            table.Date = sheet['B'+str(i + 1)].value
            table.Name = sheet['C'+str(i + 1)].value
            table.Number = sheet['D'+str(i + 1)].value
            table.Faks = sheet['E' + str(i + 1)].value
            table.Adress = sheet['F'+str(i + 1)].value
            table.Manager = sheet['I' + str(i + 1)].value
            table.Oblast = sheet['J' + str(i + 1)].value.replace('Иркутская', 'Иркутская область')
            table.Rayon = sheet['K' + str(i + 1)].value.replace('Иркутск', 'Иркутский')
            table.Source = sheet['L' + str(i + 1)].value
            table.Source2 = sheet['M' + str(i + 1)].value
            table.Segment = sheet['N' + str(i + 1)].value
            table.Segment2 = sheet['O' + str(i + 1)].value
            table.Status = sheet['P' + str(i + 1)].value
            table.Description = sheet['Q' + str(i + 1)].value


            if (sheet['A'+str(i + 2)].value == 'История'):
                result = get_clients_notes(sheet=sheet, parent_table=table, col=i + 2, client_id=client_id)
                table.Notes = result['notes']
                if str(result['contacts']) != '[]':
                    table.Contacts = result['contacts']

            if (sheet['A' + str(i + 2)].value == 'Контактные лица'):
                table.Contacts = get_clients_contacts(sheet=sheet, parent_table=table, col=i + 2, client_id=client_id)

            db.session.add(table)
            db.session.commit()
            client_id = client_id + 1

def add_client_from_xlsx():
    print(1)
    sheet = openpyxl.load_workbook('app/files/book2Irkutsk.xlsx').active
    table = models.Client
    get_client_from_xlsx(col=1, sheet=sheet, parent_table=table)

# ------------------------------------------------------Provider--------------------------------------------------------

def get_providers_from_xlsx(col, sheet, parent_table):
    for i in range(14, 18):
        table = parent_table()
        table.Name = sheet['A'+str(i)].value
        table.Adress = sheet['B'+str(i)].value
        table.Raps = sheet['E'+str(i)].value
        table.Train = sheet['I'+str(i)].value
        table.Vets = sheet['J'+str(i)].value
        db.session.add(table)

    db.session.commit()



def add_provider_from_xlsx():
    sheet = openpyxl.load_workbook('app/files/Providers.xlsx').active
    table = models.Provider
    get_providers_from_xlsx(col=1, sheet=sheet, parent_table=table)


# ------------------------------------------------------Provider--------------------------------------------------------
def get_items_from_xlsx(col, sheet, parent_table, child_table):
    table = child_table()
    table.Name = sheet['B' + str(col)].value
    table.Creator = sheet['C' + str(col)].value
    table.Bags = sheet['D' + str(col)].value
    table.MKR = sheet['E' + str(col)].value
    table.Pile = sheet['F' + str(col)].value
    table.Cost = sheet['G' + str(col)].value
    table.NDS = sheet['H' + str(col)].value
    table.Group = parent_table
    parent_table.Item.append(table)
    col = col + 1

    while not sheet['A'+str(col)].value:
        table = child_table()
        table.Name = sheet['B'+str(col)].value
        table.Creator = sheet['C'+str(col)].value
        table.Bags = sheet['D'+str(col)].value
        table.MKR = sheet['E'+str(col)].value
        table.Pile = sheet['F'+str(col)].value
        table.Cost = sheet['G'+str(col)].value
        table.NDS = sheet['H'+str(col)].value
        table.Group = parent_table
        parent_table.Item.append(table)
        if col == sheet.max_row:
            break
        col = col + 1

    return parent_table.Item


def get_groups_from_xlsx(col, sheet, parent_table, child_table):
    for i in range(col+2, sheet.max_row):
        if not sheet['A'+str(i)].value:
            continue

        groupTable = parent_table()
        groupTable.Group = sheet['A'+str(i)].value
        groupTable.Item = get_items_from_xlsx(col=i, sheet=sheet, parent_table=groupTable, child_table=child_table)
        db.session.add(groupTable)
        db.session.commit()



def add_items_from_xlsx():
    sheet = openpyxl.load_workbook('app/files/Items.xlsx').active
    table = models.Item
    get_groups_from_xlsx(col=1, sheet=sheet, parent_table=models.Item_groups, child_table=table)


# ----------------------------------------------------Templates---------------------------------------------------------
def addTemplate(name):
    table = models.Template()
    table.name = name

    db.session.add(table)
    db.session.commit()

# ----------------------------------------------------Testing_Room------------------------------------------------------
def table_to_json(query):
    result = []
    for i in query:
        subres = i.__dict__
        if '_sa_instance_state' in subres:
            subres.pop('_sa_instance_state', None)
        if 'Date' in subres:
            if subres['Date'] != None:
                subres['Date'] = subres['Date'].strftime("%m/%d/%Y, %H:%M:%S")

        result.append(subres)
    return json.dumps(result)

# -------------------------------------------------------Usage----------------------------------------------------------

# add_client_from_xlsx()
add_user(login='Dialjini', password='DaniLKA210', email='kustovdanil2@gmail.com', role='admin', name='Василий Пупкин')
add_user(login='admin', password='qwerzy132', email='лысый', role='admin', name='Василий Пупкин')
# add_provider_from_xlsx()
# addTemplate('Договор')
# addTemplate('Заявка')
add_roles()
# db.session.delete(models.Account.query.filter_by(id=1).first())
db.session.commit()
